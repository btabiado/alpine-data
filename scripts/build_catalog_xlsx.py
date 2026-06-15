#!/usr/bin/env python3
"""
build_catalog_xlsx.py — render the Data Sources Catalog as a downloadable
Excel workbook (health/api-catalog.xlsx), the spreadsheet sibling of the
landscape PDF.

Reads:
  health/api_catalog.json     the 2,773-entry catalog (categories -> entries)
  health/catalog_health.json  link-health output from check_catalog_links.py
                              (used to fill the Freshness/Health column)

Writes:
  health/api-catalog.xlsx     Summary + "All Entries" + one tab per category

The Freshness/Health column is best-effort: catalog_health.json only persists
the DEAD and UNREACHABLE URLs individually (ok/gated are aggregate counts), so
every non-flagged row shows "OK / not flagged". Dead rows are shaded red,
unreachable orange. Run after check_catalog_links.py so the two stay in sync.

Usage:
    python scripts/build_catalog_xlsx.py [--catalog PATH] [--health PATH] [--out PATH]
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (short json key, friendly column header)
COLS = [
    ("co", "Company / Provider"),
    ("n", "Service / API Name"),
    ("category", "Category"),
    ("b", "What it is"),
    ("f", "Free tier / terms"),
    ("p", "Price"),
    ("s", "Status"),
    ("health", "Freshness / Health"),
    ("fr", "Free? (1/0)"),
    ("kl", "Keyless? (1/0)"),
    ("kr", "Key required? (1/0)"),
    ("cf", "Coverage/Conf."),
    ("u", "URL"),
]
WIDTHS = [26, 34, 30, 52, 30, 16, 12, 20, 9, 9, 11, 12, 46]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
DEAD_FILL = PatternFill("solid", fgColor="F4CCCC")
UNR_FILL = PatternFill("solid", fgColor="FCE5CD")
HEALTH_IDX = [k for k, _ in COLS].index("health") + 1  # 1-based column


def _norm(u: str) -> str:
    return (u or "").strip().rstrip("/").lower()


def build_status_map(health: dict) -> callable:
    dead = {_norm(x["url"]): f"Dead ({x['code']})" for x in health.get("dead", [])}
    unr = {_norm(x["url"]): "Unreachable" for x in health.get("unreachable", [])}

    def status_for(url: str) -> str:
        n = _norm(url)
        return dead.get(n) or unr.get(n) or "OK / not flagged"

    return status_for


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_rows(ws, rows: list[dict]) -> None:
    for r in rows:
        ws.append([r.get(k, "") for k, _ in COLS])
        val = ws.cell(ws.max_row, HEALTH_IDX).value
        if val and val.startswith("Dead"):
            for cell in ws[ws.max_row]:
                cell.fill = DEAD_FILL
        elif val == "Unreachable":
            for cell in ws[ws.max_row]:
                cell.fill = UNR_FILL


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--catalog", default="health/api_catalog.json")
    ap.add_argument("--health", default="health/catalog_health.json")
    ap.add_argument("--out", default="health/api-catalog.xlsx")
    args = ap.parse_args()

    cat = json.loads(Path(args.catalog).read_text())
    meta = cat["meta"]
    health = json.loads(Path(args.health).read_text())
    status_for = build_status_map(health)

    rows: list[dict] = []
    for c in cat["categories"]:
        for e in c.get("entries", []):
            r = {"category": c["category"], "health": status_for(e.get("u", ""))}
            r.update(e)
            rows.append(r)

    headers = [c[1] for c in COLS]
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append([meta.get("title", "Data Sources Catalog")])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for k, v in (
        ("Total entries", meta.get("total")),
        ("Free entries", meta.get("free_count")),
        ("Available (net-new)", meta.get("available_count")),
        ("Categories", len(cat["categories"])),
    ):
        ws.append([k, v])
    ws.append([])
    ws.append(["Health probe (catalog_health.json)", ""])
    ws.append(["  Checked at", health.get("checked_at")])
    ws.append(["  URLs probed", health.get("total_urls")])
    for k, v in health.get("counts", {}).items():
        ws.append([f"  {k}", v])
    ws.append([])
    ws.append(["Note", meta.get("note", "")])
    ws.append([
        "Health note",
        "Only dead/unreachable URLs are flagged per-row; ok/gated are aggregate "
        "counts, so non-flagged rows show 'OK / not flagged'. Remaining "
        "unreachable are usually datacenter-IP bot walls, not genuine downtime.",
    ])
    ws.append([])
    ws.append(["Category", "Count"])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 2).font = Font(bold=True)
    for c in cat["categories"]:
        ws.append([c["category"], c["count"]])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42

    # --- All Entries ---
    wsa = wb.create_sheet("All Entries")
    wsa.append(headers)
    _style_header(wsa)
    _write_rows(wsa, rows)

    # --- One tab per category (sheet names: <=31 chars, sanitized, unique) ---
    used: set[str] = set()

    def sheet_name(name: str) -> str:
        s = re.sub(r"[\\/:*?\[\]]", "-", name)[:31].strip()
        base, i = s, 2
        while s.lower() in used or not s:
            suf = f" ({i})"
            s = base[: 31 - len(suf)] + suf
            i += 1
        used.add(s.lower())
        return s

    for c in cat["categories"]:
        ws = wb.create_sheet(sheet_name(c["category"]))
        ws.append(headers)
        _style_header(ws)
        _write_rows(ws, [r for r in rows if r["category"] == c["category"]])

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    tally = dict(Counter(r["health"] for r in rows))
    print(f"Wrote {args.out}: {len(rows)} rows, {len(wb.sheetnames)} sheets")
    print(f"Health tally: {tally}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
